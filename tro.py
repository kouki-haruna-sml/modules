import pandas as pd
import numpy as np
import datetime
import itertools
import collections
import time
import openpyxl as px
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from tqdm import tqdm_notebook as tqdm
import modules.query as qr

'''
########################
アウトプット計算用関数
########################
'''
#DBからターゲットパネラーを抽出
def get_panelers(title, date):
    #SQL
    df = qr.queryRedShift("select * from program_viewers\
                                            where prog_id in (select prog_id from programs\
                                                                            where title like {0}\
                                                                            and date = {1}\
                                                                            and channel_id in (3, 4, 5, 6, 7))\
                                         　　".format(title, date))
    df2 = qr.queryRedShift("select (ended_at - started_at) as airtime\
                                               from programs\
                                               where title like {0}\
                                               and date = {1}\
                                               and channel_id in (3, 4, 5, 6, 7)"\
                                               .format(title, date))

    if len(df) == 0:
        print("program not found")
        return

    #秒換算
    time = df2["airtime"][0]
    seconds = time.total_seconds()

    #パネラー絞り込み
    result = []
    for i in range(len(df)):
        if df["viewing_seconds"][i] >= (seconds / 3):
            result.append(df["paneler_id"][i])

    return result


#有効パネラーに限定
def filtering_valid(original_panelers, valid_panelers):
    ls = []
    for i in original_panelers:
        if i in valid_panelers:
            ls.append(i)
    return ls

#各番組のパネラー（有効に限らない）
def each_program_panelers_novalid(program_ls):
    result = []
    for i in tqdm(range(len(program_ls))):
        each_program_panelers = get_panelers(program_ls[i][0], program_ls[i][1])
        result.append(each_program_panelers)
    return result

#各番組のパネラー（有効に限る）
def each_program_panelers(program_ls, valid_panelers):
    result = []
    for i in tqdm(range(len(program_ls))):
        each_program_panelers = get_panelers(program_ls[i][0], program_ls[i][1])
        panelers = filtering_valid(each_program_panelers, valid_panelers)
        result.append(panelers)
    return result

#フォーメーション作成
def combinations(total_num, new_num): #pCq, p=total_num, q=new_num
    ls = []
    num_ls = range(total_num)
    for i in tqdm(itertools.combinations(num_ls, r=new_num)):
        ls.append(i)
    return ls

#番組毎リーチの計算
def each_program_reach(each_program_panelers_valid, valid_panelers, program_ls):
    result = {}

    for i in tqdm(range(len(each_program_panelers_valid))):
        rate = len(each_program_panelers_valid[i]) / len(valid_panelers)
        rate = round(rate * 100, 2)
        result[program_ls[i][0]] = rate

    return result

#各フォーメーションのパネラー
def each_combination_panelers(combi, each_program_panelers_valid):
    result = []
    #コンビを順番に抽出
    for i in tqdm(combi):
        combi_panelers = []
        combination = i
        #コンビの中身を順番に抽出
        for j in combination:
            combi_panelers += each_program_panelers_valid[j]
        counter = collections.Counter(combi_panelers)
        result.append(counter)
    return result

def original_total_and_unique_reach(original_num, each_program_panelers_valid, each_combi_panelers):
    result = {}

    #オリジナルのトータルリーチの計算
    panelers = []
    for i in range(14):
        for j in each_program_panelers_valid[i]:
            if j not in panelers:
                panelers.append(j)
    original_total_reach = (len(panelers) / len(valid_panelers)) * 100
    original_total_reach_rounded = round(original_total_reach, 2)

    #オリジナルのリーチドパネラー
    original_reached = {}
    for i in range(original_num):
        for j in each_program_panelers_valid[i]:
            if j in original_reached.keys():
                original_reached[j] += 1
            else:
                original_reached[j] = 1

    #オリジナルのユニークリーチの計算
    for i in range(original_num):
        tp = original_reached.copy()

        for j in each_program_panelers_valid[i]:
            if tp[j] == 1:
                tp.pop(j)

        rate = (original_total_reach - (len(tp) / len(valid_panelers))*100)
        rate = round(rate, 2)
        result[i] = rate

    return original_total_reach_rounded, result


#トータルリーチの計算
def all_total_reach(each_combi_panelers, valid_panelers):
    result = {}
    for i, j in tqdm(enumerate(range(len(each_combi_panelers)))):
        rate = len(each_combi_panelers[i]) / len(valid_panelers)
        result[i] = rate

    result_sorted = sorted(result.items(), key=lambda x:x[1], reverse=True)
    return result_sorted

#top10000に絞り込み
def top10000(total_reach, each_combi_panelers):
    total_reach_top10000 = total_reach[:1000]
    combi_top10000 = []
    reach_ls_top10000 = []
    each_combi_panelers_top10000 = []

    for i in total_reach_top10000:
        combi_top10000.append(i[0])
        reach_ls_top10000.append(i[1])
        each_combi_panelers_top10000.append(each_combi_panelers[i[0]])

    return combi_top10000, reach_ls_top10000, each_combi_panelers_top10000

#各番組毎のユニークリーチ
def unique_reach(combi_num, combi, each_program_panelers_valid, each_combi_panelers, total_reach):
    result = {}
    total_reach_dic = dict(total_reach)
    original_rate = total_reach_dic[combi_num]

    for i in combi[combi_num]:
        tp = each_combi_panelers[combi_num].copy()

        for j in each_program_panelers_valid[i]:
            if tp[j] == 1:
                tp.pop(j)

        rate = (original_rate - (len(tp) / len(valid_panelers))) * 100
        rate = round(rate, 2)
        result[i] = rate

    return result

#ユニークリーチ計算
def all_unique_reach(combi_top10000, combi, each_program_panelers_valid, each_combi_panelers, total_reach):
    result = []
    for i in tqdm(combi_top10000):
        rate_ls = unique_reach(i, combi, each_program_panelers_valid, each_combi_panelers, total_reach)
        result.append(rate_ls)
    return result

#全結果計算
def get_all(total_num, original_num, new_num, program_ls, valid_panelers):
    print("calculating each_program_panelers")
    each_program_panelers_valid = each_program_panelers(program_ls, valid_panelers)

    print("creating formations")
    combi = combinations(total_num, new_num)

    print("calculating each_program_reach")
    each_prog_reach = each_program_reach(each_program_panelers_valid, valid_panelers, program_ls)

    print("########番組毎リーチ集計完了##############")
    print("")
    print("calculating each_combination_panelers")
    each_combi_panelers = each_combination_panelers(combi, each_program_panelers_valid)

    print("calculating original total_reach and unique_reach")
    print("")
    original_rate_ls = original_total_and_unique_reach(original_num, each_program_panelers_valid, each_combi_panelers)
    original_total_reach = original_rate_ls[0]
    original_unique_reach = original_rate_ls[1]

    print("############オリジナルのトータルリーチとユニークリーチ集計完了#############")

    print("calculating total_reach")
    total_reach = all_total_reach(each_combi_panelers, valid_panelers)

    print("#########トータルリーチ集計完了##############")
    print("")

    print("filtering top10000")
    print("")
    top = top10000(total_reach, each_combi_panelers)
    combi_top10000 = top[0]

    print("calculating unique_reach")

    ureach = all_unique_reach(combi_top10000, combi, each_program_panelers_valid, each_combi_panelers, total_reach)

    print("############上位1000ユニークリーチ集計完了#################")

    return each_prog_reach, original_total_reach, original_unique_reach, total_reach, ureach


'''
#######################
excel出力用関数
#######################
'''
def program_info(program_ls):
    #放送局、放送曜日、開始時刻、終了時刻、番組名
    result = []

    for i in tqdm(program_ls):
        title = i[0]
        date = i[1]
        sql = "select c.code_name, p.started_at, p.ended_at, p.title from programs p\
                    inner join channels c on p.channel_id = c.id\
                    where p.title like {0}\
                    and p.date = {1}\
                    and c.region_id = 1\
                    ;".format(title, date)
        df = qr.queryRedShift(sql)

        code_name = df["code_name"][0]
        weekday = df["started_at"][0].weekday_name
        started_at = df["started_at"][0].strftime("%H:%M:%S")
        ended_at = df["ended_at"][0].strftime("%H:%M:%S")
        title = df["title"][0]

        wday_dic = {"Monday": "月", "Tuesday": "火", "Wednesday": "水", "Thursday": "木", "Friday": "金", "Saturday": "土", "Sunday": "日"}
        weekday = wday_dic[weekday]

        result.append([code_name, weekday, started_at, ended_at, title])

    return result

def make_excel(total_num, original_num, new_num, prog_info, result):
    wb = px.Workbook()
    ws = wb.active

    #左ページカラム名の設定
    columns = ["放送局", "放送曜日", "通常開始時刻", "通常終了時刻", "番組名", "番組毎リーチ", "トータルリーチ"]
    for i in range(1, 1+len(columns)):
        ws.cell(row=3, column=i).value = columns[i-1]
        ws.cell(row=3, column=i).fill = PatternFill(patternType="solid", fgColor="d3d3d3")

    #左ページ各カラムのバリュー
    for i in range(len(prog_info[0])):
        values = [x[i] for x in prog_info]
        for j in range(4, 4+len(values)):
            ws.cell(row=j, column=i+1).value = values[j-4]

    #番組毎リーチ埋める
    for i, j in enumerate(result[0].values()):
        ws.cell(row=4+i, column=6).value = j

    #「提供中のみ」のトータルリーチ
    ws.cell(row=3, column=8).value = result[1]

    #「提供中のみ」のユニークリーチ
    for i, j in enumerate(result[2].values()):
            ws.cell(row=4+i, column=8).value = j
            #赤色塗り
            if j < 1:
                ws.cell(row=4+i, column=8).fill = PatternFill(patternType="solid", fgColor="FBE0E0")
            elif j < 2:
                ws.cell(row=4+i, column=8).fill = PatternFill(patternType="solid", fgColor="F5A9A9")
            else:
                ws.cell(row=4+i, column=8).fill = PatternFill(patternType="solid", fgColor="FA5858")

    #トータルリーチ
    for i, j in enumerate(dict(result[3][:1000]).values()):
        rate = j * 100
        rate = round(rate, 2)
        ws.cell(row=3, column=9+i).value = rate

    #番組毎ユニークリーチ
    for i in range(len(result[4])):
        dic = result[4][i]
        for j, k in enumerate(dic.keys()):
            ws.cell(row=4+k, column=9+i).value = dic[k]
            #赤色塗り
            if dic[k] < 1:
                ws.cell(row=4+k, column=9+i).fill = PatternFill(patternType="solid", fgColor="FBE0E0")
            elif dic[k] < 2:
                ws.cell(row=4+k, column=9+i).fill = PatternFill(patternType="solid", fgColor="F5A9A9")
            else:
                ws.cell(row=4+k, column=9+i).fill = PatternFill(patternType="solid", fgColor="FA5858")

    #トータルリーチ順位行バリュー、色塗り
    ws.cell(row=2, column=8).value = "提供中のみ"

    for i in range(1001):
        ws.cell(row=2, column=8+i).fill = PatternFill(patternType="solid", fgColor="d3d3d3")
        if not i == 1000:
            ws.cell(row=2, column=9+i).value = "{}位".format(i+1)

    #左上のタイトル
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=7)

    #E列色塗り
    for i in range(4, 4+original_num):
        ws.cell(row=i, column=5).fill = PatternFill(patternType="solid", fgColor="D3EDFB")

    for i in range(4+original_num, 4+total_num):
        ws.cell(row=i, column=5).fill = PatternFill(patternType="solid", fgColor="FFFCDB")

    #G列「ユニークリーチ」
    ws.cell(row=13, column=7).value = "ユニークリーチ"

    #G列色塗り
    for i in range(4, 27):
        ws.cell(row=i, column=7).fill = PatternFill(patternType="solid", fgColor="d3d3d3")

    #枠線設定
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)

    for row in ws:
        for cell in row:
            ws[cell.coordinate].border = border

    #G列結合
    ws.merge_cells(start_row=4, start_column=7, end_row=26, end_column=7)

    #セルの幅を自動設定
    for col in ws.columns:
        max_length = 0
        column = px.utils.get_column_letter(col[0].column)

        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length) * 2
        ws.column_dimensions[column].width = adjusted_width

    #フォントを「メイリオ」に指定
    font = px.styles.Font(name='メイリオ')
    for row in ws:
        for cell in row:
            ws[cell.coordinate].font = font

    wb.save("example.xlsx")
